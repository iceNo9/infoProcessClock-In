import datetime
from daycheck import DayCheck
import traceback
from openpyxl.styles import PatternFill

def calculate_hour_difference(start: datetime.datetime, end: datetime.datetime) -> float:
    """
    计算两个时间点之间的小时差，按照以下规则：
    - 差值 >= 45分钟视为1小时
    - 差值 >= 15分钟且 < 45分钟视为0.5小时
    - 差值 < 15分钟视为0小时

    参数：
    start (datetime.datetime): 开始时间，datetime 对象
    end (datetime.datetime): 结束时间，datetime 对象
    
    返回：
    float: 两个时间点之间的小时差，精确到 0.5 小时
    """
    
    # 计算两个时间点之间的秒数差
    time_diff = (end - start).total_seconds()
    
    # 将秒数转换为小时
    hours_diff = time_diff / 3600  # 1小时 = 3600秒
    
    # 提取小时和分钟部分
    full_hours = int(hours_diff)  # 完整小时
    minutes_diff = (hours_diff - full_hours) * 60  # 计算剩余的分钟部分
    
    # 计算分钟部分
    if minutes_diff >= 45:
        return full_hours + 1.0  # 超过45分钟，视为1小时
    elif minutes_diff >= 15:
        return full_hours + 0.5  # 超过15分钟但不到45分钟，视为0.5小时
    else:
        return full_hours  # 小于15分钟，视为0小时
    
# 获取中文星期
def get_weekday_chinese(date_str):
    date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    weekday_num = date.weekday()  # 返回0=Monday, 1=Tuesday, ..., 6=Sunday
    weekday_chinese = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    return weekday_chinese[weekday_num]

class NonWorkdayAttendance:
    def __init__(self):
        self.status = None  # 状态
        self.work_start_time = None  # 上班时间
        self.work_end_time = None  # 下班时间
        self.overtime_hours = 0


    def set_status(self, status, work_start_time=None, work_end_time=None):
        """设置状态、上班时间和下班时间"""
        self.status = status
        self.work_start_time = work_start_time
        self.work_end_time = work_end_time

class WorkdayAttendance:
    # 定义一个类变量，用于保存整个周期（如一个月）内的总加班时长
    total_overtime_hours = 0

    def __init__(self, date: datetime.date):
        self.date = date
        self.morning_in = {"status": None, "time": None}
        self.morning_out = {"status": None, "time": None}
        self.afternoon_in = {"status": None, "time": None}
        self.afternoon_out = {"status": None, "time": None}
        self.overtime_in = {"status": None, "time": None}
        self.overtime_out = {"status": None, "time": None}        
        self.overtime_hours = 0

    def set_status(self, period, status, time):
        """设置打卡状态和打卡时间"""
        if period == "morning_in":
            self.morning_in["status"] = status
            self.morning_in["time"] = time
        elif period == "morning_out":
            self.morning_out["status"] = status
            self.morning_out["time"] = time
        elif period == "afternoon_in":
            self.afternoon_in["status"] = status
            self.afternoon_in["time"] = time
        elif period == "afternoon_out":
            self.afternoon_out["status"] = status
            self.afternoon_out["time"] = time
        elif period == "overtime_in":
            self.overtime_in["status"] = status
            self.overtime_in["time"] = time
        elif period == "overtime_out":
            self.overtime_out["status"] = status
            self.overtime_out["time"] = time

        def add_overtime(self, overtime_hours: float):
            """
            增加加班时长到该天的加班时长
            """
            self.overtime_hours += overtime_hours
            WorkdayAttendance.total_overtime_hours += overtime_hours  # 更新全局总加班时长

class AttendanceManager:
    def __init__(self, year, is_flexible=True):
        self.year = year
        self.is_flexible = is_flexible  # 是否开启弹性工作制
        self.day_check = DayCheck(year)  # 初始化日期判断类

    def process_attendance(self, attendance_data):
        """处理传入的考勤数据字典"""
        result = {}
        for date_str, punches in attendance_data.items():
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            result[date_str] = self.check_in_out(date, punches)
        return result

    def check_in_out(self, date: datetime.date, punches: list):
        """检查打卡数据"""

        # 将打卡时间转换为datetime对象
        punches = [datetime.datetime.strptime(p, "%Y-%m-%d %H:%M:%S") for p in punches]
        
        # 根据日期判断类型
        day_type = self.day_check.get_day_type(date)

        # 检查并处理工作日考勤
        if day_type == "workday":
            return self.handle_workday(date, punches)
        elif day_type == "restday":
            return self.handle_restday(punches)
        elif day_type == "holiday":
            return self.handle_holiday(punches)
        else:
            return "日期类型未知"

    
    def handle_workday(self, date: datetime.date, punches: list):
        """处理工作日考勤"""
        # 定义固定的基础时间
        AM_WORK_START = datetime.time(8, 30)  # 上午上班时间
        AM_WORK_END = datetime.time(12, 10)   # 上午下班时间
        PM_WORK_START = datetime.time(13, 40)  # 下午上班时间
        PM_WORK_END = datetime.time(18, 0)    # 下午下班时间
        AM_PM_LINE_TIME = datetime.time(13, 00)  # >=此时间为下午

        # 时间间隔类型（更改为timedelta）
        FLEXIBLE_TIME = datetime.timedelta(hours=0, minutes=30)    # 弹性时间
        OVERTIME = datetime.timedelta(hours=0, minutes=45)  # 加班时间
        EAT_TIME = datetime.timedelta(hours=0, minutes=30)  # 休息时间

        # 确保所有时间是根据当前日期生成的
        am_start_time = datetime.datetime.combine(date, AM_WORK_START)
        am_end_time = datetime.datetime.combine(date, AM_WORK_END)
        pm_start_time = datetime.datetime.combine(date, PM_WORK_START)
        pm_end_time = datetime.datetime.combine(date, PM_WORK_END)
        am_pm_line_time = datetime.datetime.combine(date, AM_PM_LINE_TIME)

        # 创建工作日考勤记录
        workday_attendance = WorkdayAttendance(date)

        # 计算是否需要延长下午下班时间
        afternoon_extension = datetime.timedelta(hours=0, minutes=0)

        # 午间打卡信息列表
        middle_list = []

        # 遍历打卡时间列表，依次进行状态判断
        # 使用迭代器
        punch_iterator = iter(punches)
        while True:
            try:                
                punch = next(punch_iterator)  # 获取下一个打卡时间

                # 处理[0,8:30]，上午上班判断
                if am_start_time >= punch:            
                    if workday_attendance.morning_in["status"] is None:
                        workday_attendance.set_status("morning_in", "正常", punch)

                # 处理(8:30,9:00]，上午上班判断
                elif am_start_time+FLEXIBLE_TIME >= punch:
                    if self.is_flexible:
                        if workday_attendance.morning_in["status"] is None:
                            workday_attendance.set_status("morning_in", "正常", punch)
                            afternoon_extension = punch - am_start_time

                # 处理(9:00,12:10)，上午上班判断&上午下班判断
                elif am_end_time > punch:
                    if workday_attendance.morning_in["status"] is None:
                        workday_attendance.set_status("morning_in", "迟到", punch)
                    else:
                        workday_attendance.set_status("morning_out", "早退", punch)

                # 处理[12:10,13:40]
                elif pm_start_time >= punch:
                    if workday_attendance.morning_in["status"] is None: # 上午上班缺卡判断
                        workday_attendance.set_status("morning_in", "缺卡", None)                    
                    middle_list.append(punch)

                # 处理(13:40,18:00+afternoon_extension)
                elif (pm_end_time+afternoon_extension) > punch:
                    if len(middle_list) == 2:
                        workday_attendance.set_status("morning_out", "正常", middle_list[0])
                        workday_attendance.set_status("afternoon_in", "正常", middle_list[1])
                        workday_attendance.set_status("afternoon_out", "早退", punch)
                    elif len(middle_list) == 1:
                        if am_pm_line_time > middle_list[0]: #视为上午
                            workday_attendance.set_status("morning_out", "正常", middle_list[0])
                            workday_attendance.set_status("afternoon_in", "迟到", punch)
                        else:
                            workday_attendance.set_status("morning_out", "缺卡", None)
                            workday_attendance.set_status("afternoon_in", "正常", middle_list[0])
                            workday_attendance.set_status("afternoon_out", "早退", punch)                        
                    elif len(middle_list) == 0:
                        workday_attendance.set_status("morning_out", "缺卡", None)
                        workday_attendance.set_status("afternoon_in", "缺卡", None)
                        workday_attendance.set_status("afternoon_out", "早退", punch)  
                    else:
                        #不该有其他数目
                        print(f"午间存在3次及以上打卡异常:{middle_list}")

                # 处理[18:00+afternoon_extension,18:00+afternoon_extension+pm_overtime) 下午下班时间
                elif (pm_end_time+afternoon_extension+OVERTIME) > punch:
                    workday_attendance.set_status("afternoon_out", "正常", punch)  

                    if workday_attendance.morning_out["status"] is None or workday_attendance.afternoon_in["status"] is None:
                        if len(middle_list) == 2:       
                            workday_attendance.set_status("morning_out", "正常", middle_list[0])
                            workday_attendance.set_status("afternoon_in", "正常", middle_list[1])      
                        elif len(middle_list) == 1:
                            if am_pm_line_time > middle_list[0]: #视为上午
                                workday_attendance.set_status("morning_out", "正常", middle_list[0])
                                workday_attendance.set_status("afternoon_in", "缺卡", None)
                                
                            else:
                                workday_attendance.set_status("morning_out", "缺卡", None)
                                workday_attendance.set_status("afternoon_in", "正常", middle_list[0])
                        elif len(middle_list) == 0:
                            workday_attendance.set_status("morning_out", "缺卡", None)
                            workday_attendance.set_status("afternoon_in", "缺卡", None)
                        else:
                            #不该有其他数目
                            print(f"午间存在3次及以上打卡异常:{middle_list}")

                # 处理加班
                else:
                    workday_attendance.set_status("overtime_in", "正常", pm_end_time+afternoon_extension+EAT_TIME)
                    workday_attendance.set_status("overtime_out", "正常", punch)
                    workday_attendance.set_status("afternoon_out", "加班", punch)
                    workday_attendance.overtime_hours=calculate_hour_difference(workday_attendance.overtime_in["time"],workday_attendance.overtime_out["time"])

                    if workday_attendance.morning_out["status"] is None or workday_attendance.afternoon_in["status"] is None:
                        if len(middle_list) == 2:       
                            workday_attendance.set_status("morning_out", "正常", middle_list[0])
                            workday_attendance.set_status("afternoon_in", "正常", middle_list[1])      
                        elif len(middle_list) == 1:
                            if am_pm_line_time > middle_list[0]: #视为上午
                                workday_attendance.set_status("morning_out", "正常", middle_list[0])
                                workday_attendance.set_status("afternoon_in", "缺卡", None)
                                
                            else:
                                workday_attendance.set_status("morning_out", "缺卡", None)
                                workday_attendance.set_status("afternoon_in", "正常", middle_list[0])
                        elif len(middle_list) == 0:
                            workday_attendance.set_status("morning_out", "缺卡", None)
                            workday_attendance.set_status("afternoon_in", "缺卡", None)
                        else:
                            #不该有其他数目
                            print(f"午间存在3次及以上打卡异常:{middle_list}")

            except StopIteration:
                break  # 当迭代器遍历完毕时跳出循环     
            except Exception as e:
                # 打印详细的错误信息，包括行号
                print(f"An error occurred: {e}")
                print("Detailed traceback:")
                traceback.print_exc()  # 打印错误的详细信息，包括行号

        if 4 > len(punches):
            if workday_attendance.morning_in["status"] is None:
                workday_attendance.set_status("morning_in", "缺卡", None)

            if workday_attendance.morning_out["status"] is None:
                workday_attendance.set_status("morning_out", "缺卡", None)

            if workday_attendance.afternoon_in["status"] is None:
                workday_attendance.set_status("afternoon_in", "缺卡", None)

            if workday_attendance.afternoon_out["status"] is None:
                workday_attendance.set_status("afternoon_out", "缺卡", None)


        return workday_attendance

    def handle_restday(self, punches: list):
        """处理休息日"""
        non_workday_attendance = NonWorkdayAttendance()

        if len(punches) == 2:  # 如果有上下班时间，则视为加班
            non_workday_attendance.set_status("公休加班", punches[0], punches[1])
            non_workday_attendance.overtime_hours=calculate_hour_difference(punches[0], punches[1])
        elif len(punches) == 1:  # 如果只有一个打卡时间，则视为缺卡
            non_workday_attendance.set_status("缺卡", punches[0], None)
        else:  # 如果没有任何打卡记录，则视为正常
            non_workday_attendance.set_status("正常")

        return non_workday_attendance


    def handle_holiday(self, punches: list):
        """处理节假日"""
        non_workday_attendance = NonWorkdayAttendance()

        if len(punches) == 2:  # 如果有上下班时间，则视为加班
            non_workday_attendance.set_status("节日加班", punches[0], punches[1])
            non_workday_attendance.overtime_hours=calculate_hour_difference(punches[0], punches[1])
        elif len(punches) == 1:  # 如果只有一个打卡时间，则视为缺卡
            non_workday_attendance.set_status("缺卡", punches[0], None)
        else:  # 如果没有任何打卡记录，则视为正常
            non_workday_attendance.set_status("正常")

        return non_workday_attendance

    def process_month(self, month, attendance_data):
        """处理指定年份和月份的考勤情况"""
        # 获取该月的所有日期
        month_dates = [datetime.date(self.year, month, day) for day in range(1, 32)
                    if (datetime.date(self.year, month, day)).month == month]

        results = {}

        for date in month_dates:
            date_str = date.strftime("%Y-%m-%d")
            if date_str in attendance_data:
                punches = attendance_data[date_str]
                result = self.check_in_out(date, punches)
                results[date_str] = result

                # # 打印详细信息
                # print(f"{date_str}:")
                # if isinstance(result, WorkdayAttendance):
                #     print(f"  上午上班: {result.morning_in['status']} (打卡时间: {result.morning_in['time']})")
                #     print(f"  上午下班: {result.morning_out['status']} (打卡时间: {result.morning_out['time']})")
                #     print(f"  下午上班: {result.afternoon_in['status']} (打卡时间: {result.afternoon_in['time']})")
                #     print(f"  下午下班: {result.afternoon_out['status']} (打卡时间: {result.afternoon_out['time']})")
                #     # 打印加班信息
                #     print(f"  加班开始: {result.overtime_in['status']} (打卡时间: {result.overtime_in['time']})")
                #     print(f"  加班结束: {result.overtime_out['status']} (打卡时间: {result.overtime_out['time']})")
                #     print(f"  加班时长: {result.overtime_hours} 小时")
                # elif isinstance(result, NonWorkdayAttendance):
                #     print(f"  状态: {result.status}")
                #     if result.work_start_time:
                #         print(f"  上班时间: {result.work_start_time}")
                #     if result.work_end_time:
                #         print(f"  下班时间: {result.work_end_time}")
            else:
                day_type = self.day_check.get_day_type(date)
                if day_type == "workday":
                    results[date_str] = "缺勤"
                    # print(f"{date_str}: 缺勤")
                else:
                    results[date_str] = "非工作日"
                    # print(f"{date_str}: 非工作日")

        return results

    def write_attendance_to_excel2(self, wb, attendance_data):
        """
        将考勤数据写入到 Excel 工作簿的 `detail` 表中。

        参数：
        wb (openpyxl.Workbook): 一个工作簿对象
        attendance_data (dict): 包含考勤数据的字典
        """
        # 删除默认的工作表
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 创建一个工作表名为 "detail"
        ws = wb.create_sheet(title="detail")

        # 写入表头
        ws.append([
            "日期", "星期", "上午上班状态", "上午上班时间", "上午下班状态", "上午下班时间",
            "下午上班状态", "下午上班时间", "下午下班状态", "下午下班时间", 
            "加班开始状态", "加班开始时间", "加班结束状态", "加班结束时间", "加班时长"
        ])

        # 遍历考勤数据字典，逐行写入
        for date_str, data in attendance_data.items():
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            weekday = get_weekday_chinese(date_str)

            # 初始化每行的数据
            row = [date_str, weekday]

            # 处理工作日考勤数据
            if isinstance(data, WorkdayAttendance):
                # 上午上班
                row.append(data.morning_in["status"] or "缺卡")
                row.append(data.morning_in["time"] if data.morning_in["time"] else "缺卡")
                # 上午下班
                row.append(data.morning_out["status"] or "缺卡")
                row.append(data.morning_out["time"] if data.morning_out["time"] else "缺卡")
                # 下午上班
                row.append(data.afternoon_in["status"] or "缺卡")
                row.append(data.afternoon_in["time"] if data.afternoon_in["time"] else "缺卡")
                # 下午下班
                row.append(data.afternoon_out["status"] or "缺卡")
                row.append(data.afternoon_out["time"] if data.afternoon_out["time"] else "缺卡")
                # 加班
                row.append(data.overtime_in["status"] or "缺卡")
                row.append(data.overtime_in["time"] if data.overtime_in["time"] else "缺卡")
                row.append(data.overtime_out["status"] or "缺卡")
                row.append(data.overtime_out["time"] if data.overtime_out["time"] else "缺卡")
                row.append(data.overtime_hours)  # 加班时长
            elif isinstance(data, NonWorkdayAttendance):
                # 非工作日数据
                row.extend([data.status, data.work_start_time if data.work_start_time else "", 
                            data.work_end_time if data.work_end_time else ""])
                row.extend([""] * 10)  # 其他列为空

            # 写入当前行数据
            ws.append(row)

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        print("考勤数据已成功写入 Excel 文件的 detail 表中。")



    def write_attendance_to_excel2(self, wb, attendance_data):
        """
        将考勤数据写入到 Excel 工作簿的 `detail` 表中。

        参数：
        wb (openpyxl.Workbook): 一个工作簿对象
        attendance_data (dict): 包含考勤数据的字典
        """
        # 删除默认的工作表
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 创建一个工作表名为 "detail"
        ws = wb.create_sheet(title="detail")

        # 定义黄色和红色的单元格填充样式
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # 写入表头
        ws.append([
            "日期", "星期", "类型", "状态",  # 新增的“类型”和“状态”列
            "上午上班时间", "上午下班时间", "下午上班时间", "下午下班时间", 
            "加班开始时间", "加班结束时间", "加班时长"
        ])

        # 遍历考勤数据字典，逐行写入
        for date_str, data in attendance_data.items():
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            weekday = get_weekday_chinese(date_str)

            # 判断类型（工作日/非工作日）
            day_type = self.day_check.get_day_type(date)
            if "holiday" == day_type:
                day_type = "节假日"
            elif "restday" == day_type:
                day_type = "公休日"
            else:
                day_type = "工作日"
            # day_type = "工作日" if isinstance(data, WorkdayAttendance) else "非工作日"

            # 判断状态（异常、加班、正常）
            status = "正常"  # 默认是正常

            if data == "缺勤":
                status = "异常"
            
            elif isinstance(data, WorkdayAttendance):
                # 检查是否存在缺卡、迟到或早退
                if any([data.morning_in["status"] in ["缺卡", "迟到"],
                        data.morning_out["status"] in ["缺卡", "早退"],
                        data.afternoon_in["status"] in ["缺卡", "迟到"],
                        data.afternoon_out["status"] in ["缺卡", "早退"]]):
                    status = "异常"
                # 如果加班时间存在，判断为加班
                elif any([data.overtime_in["status"] != None, data.overtime_out["status"] != None]):
                    status = "加班"
            
            elif isinstance(data, NonWorkdayAttendance):
                if any([data.status in ["缺卡", "迟到"]]):
                    status = "异常"
                # 如果加班时间存在，判断为加班
                elif any([data.status != None]):
                    status = "加班"

            # 初始化每行的数据
            row = [date_str, weekday, day_type, status]

            # 处理工作日考勤数据
            if isinstance(data, WorkdayAttendance):
                # 上午上班时间
                row.append(self._get_time_or_empty(data.morning_in))
                # 上午下班时间
                row.append(self._get_time_or_empty(data.morning_out))
                # 下午上班时间
                row.append(self._get_time_or_empty(data.afternoon_in))
                # 下午下班时间
                row.append(self._get_time_or_empty(data.afternoon_out))
                # 加班开始时间
                row.append(self._get_time_or_empty(data.overtime_in))
                # 加班结束时间
                row.append(self._get_time_or_empty(data.overtime_out))
                # 加班时长
                row.append(data.overtime_hours if data.overtime_hours else "")
            elif isinstance(data, NonWorkdayAttendance):
                # 非工作日数据
                row.extend([data.status, data.work_start_time if data.work_start_time else "", 
                            data.work_end_time if data.work_end_time else ""])
                row.extend([""] * 6)  # 其他列为空

            # 写入当前行数据
            ws.append(row)

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        print("考勤数据已成功写入 Excel 文件的 detail 表中。")

    def write_attendance_to_excel3(self, wb, attendance_data):
        """
        将考勤数据写入到 Excel 工作簿的 `detail` 表中。

        参数：
        wb (openpyxl.Workbook): 一个工作簿对象
        attendance_data (dict): 包含考勤数据的字典
        """
        # 删除默认的工作表
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 创建一个工作表名为 "detail"
        ws = wb.create_sheet(title="detail")

        # 定义黄色和红色的单元格填充样式
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # 写入表头
        ws.append([
            "日期", "星期", "类型", "状态",  # 新增的“类型”和“状态”列
            "上午上班时间", "上午下班时间", "下午上班时间", "下午下班时间", 
            "加班开始时间", "加班结束时间", "加班时长"
        ])

        # 遍历考勤数据字典，逐行写入
        for date_str, data in attendance_data.items():
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            weekday = get_weekday_chinese(date_str)

            # 判断类型（工作日/非工作日）
            day_type = self.day_check.get_day_type(date)
            if "holiday" == day_type:
                day_type = "节假日"
            elif "restday" == day_type:
                day_type = "公休日"
            else:
                day_type = "工作日"

            # 判断状态（异常、加班、正常）
            status = "正常"  # 默认是正常

            if data == "缺勤":
                status = "异常"
            
            elif isinstance(data, WorkdayAttendance):
                # 检查是否存在缺卡、迟到或早退
                if any([data.morning_in["status"] in ["缺卡", "迟到"],
                        data.morning_out["status"] in ["缺卡", "早退"],
                        data.afternoon_in["status"] in ["缺卡", "迟到"],
                        data.afternoon_out["status"] in ["缺卡", "早退"]]):
                    status = "异常"
                # 如果加班时间存在，判断为加班
                elif any([data.overtime_in["status"] != None, data.overtime_out["status"] != None]):
                    status = "加班"
            
            elif isinstance(data, NonWorkdayAttendance):
                if any([data.status in ["缺卡", "迟到"]]):
                    status = "异常"
                # 如果加班时间存在，判断为加班
                elif any([data.status != None]):
                    status = "加班"

            # 初始化每行的数据
            row = [date_str, weekday, day_type, status]

            # 处理工作日考勤数据
            if isinstance(data, WorkdayAttendance):
                # 上午上班时间
                row.append(self._get_time_or_empty(data.morning_in))
                # 上午下班时间
                row.append(self._get_time_or_empty(data.morning_out))
                # 下午上班时间
                row.append(self._get_time_or_empty(data.afternoon_in))
                # 下午下班时间
                row.append(self._get_time_or_empty(data.afternoon_out))
                # 加班开始时间
                row.append(self._get_time_or_empty(data.overtime_in))
                # 加班结束时间
                row.append(self._get_time_or_empty(data.overtime_out))
                # 加班时长
                row.append(data.overtime_hours if data.overtime_hours else "")
            elif isinstance(data, NonWorkdayAttendance):
                # 非工作日数据
                row.extend([data.status, data.work_start_time if data.work_start_time else "", 
                            data.work_end_time if data.work_end_time else ""])
                row.extend([""] * 6)  # 其他列为空

            # 判断当前行的状态，如果是“加班”或“异常”，填充颜色
            if status == "加班":
                # 填充黄色
                for i in range(1, 12):  # 填充从第一个表头到最后一个表头的所有单元格（11个列）
                    ws.cell(row=ws.max_row, column=i).fill = yellow_fill
            elif status == "异常":
                # 填充红色
                for i in range(1, 12):  # 填充从第一个表头到最后一个表头的所有单元格（11个列）
                    ws.cell(row=ws.max_row, column=i).fill = red_fill

            # 写入当前行数据
            ws.append(row)

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        print("考勤数据已成功写入 Excel 文件的 detail 表中。")

    def write_attendance_to_excel(self, wb, attendance_data):
        """
        将考勤数据写入到 Excel 工作簿的 `detail` 表中。

        参数：
        wb (openpyxl.Workbook): 一个工作簿对象
        attendance_data (dict): 包含考勤数据的字典
        """
        # 删除默认的工作表
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 创建一个工作表名为 "detail"
        ws = wb.create_sheet(title="detail")

        # 定义黄色和红色的单元格填充样式
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # 写入表头
        ws.append([
            "日期", "星期", "类型", "状态",  # 新增的“类型”和“状态”列
            "上午上班时间", "上午下班时间", "下午上班时间", "下午下班时间", 
            "加班开始时间", "加班结束时间", "加班时长", "加班原因"
        ])

        # 遍历考勤数据字典，逐行写入
        for date_str, data in attendance_data.items():
            date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            weekday = get_weekday_chinese(date_str)

            # 判断类型（工作日/非工作日）
            day_type = self.day_check.get_day_type(date)
            if "holiday" == day_type:
                day_type = "节假日"
            elif "restday" == day_type:
                day_type = "公休日"
            else:
                day_type = "工作日"

            # 判断状态（异常、加班、正常）
            status = "正常"  # 默认是正常

            if data == "缺勤":
                status = "异常"
            
            elif isinstance(data, WorkdayAttendance):
                # 检查是否存在缺卡、迟到或早退
                if any([data.morning_in["status"] in ["缺卡", "迟到"],
                        data.morning_out["status"] in ["缺卡", "早退"],
                        data.afternoon_in["status"] in ["缺卡", "迟到"],
                        data.afternoon_out["status"] in ["缺卡", "早退"]]):
                    status = "异常"
                # 如果加班时间存在，判断为加班
                elif any([data.overtime_in["status"] != None, data.overtime_out["status"] != None]):
                    status = "普通加班"
            
            elif isinstance(data, NonWorkdayAttendance):
                if any([data.status in ["缺卡", "迟到"]]):
                    status = "异常"
                # 如果加班时间存在，判断为加班
                elif any([data.status == "节日加班"]):
                    status = "节日加班"
                elif any([data.status == "公休加班"]):
                    status = "公休加班"

            # 初始化每行的数据
            row = [date_str, weekday, day_type, status]

            # 处理工作日考勤数据
            if isinstance(data, WorkdayAttendance):
                # 上午上班时间
                row.append(self._get_time_or_empty(data.morning_in))
                # 上午下班时间
                row.append(self._get_time_or_empty(data.morning_out))
                # 下午上班时间
                row.append(self._get_time_or_empty(data.afternoon_in))
                # 下午下班时间
                row.append(self._get_time_or_empty(data.afternoon_out))
                # 加班开始时间
                row.append(self._get_time_or_empty(data.overtime_in))
                # 加班结束时间
                row.append(self._get_time_or_empty(data.overtime_out))
                # 加班时长
                row.append(data.overtime_hours if data.overtime_hours else "")
            elif isinstance(data, NonWorkdayAttendance):
                # 非工作日数据
                row.extend([data.work_start_time if data.work_start_time else "", 
                            data.work_end_time if data.work_end_time else ""])
                row.extend([""] * 2)  # 其他列为空
                row.extend([data.work_start_time if data.work_start_time else "", 
                            data.work_end_time if data.work_end_time else ""])
                row.extend([data.overtime_hours if data.overtime_hours else "",])

            # 写入当前行数据
            ws.append(row)

            # 获取当前行的行号（在写入后，ws.max_row 增加了1）
            current_row = ws.max_row

            # 判断当前行的状态，如果是“加班”或“异常”，填充颜色
            if "加班" in status:
                # 填充黄色
                for i in range(1, 12):  # 填充从第一个表头到最后一个表头的所有单元格（11个列）
                    ws.cell(row=current_row, column=i).fill = yellow_fill
            elif status == "异常":
                # 填充红色
                for i in range(1, 12):  # 填充从第一个表头到最后一个表头的所有单元格（11个列）
                    ws.cell(row=current_row, column=i).fill = red_fill

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    # 获取当前单元格的内容并更新最长长度
                    cell_value = str(cell.value) if cell.value else ""
                    max_length = max(max_length, len(cell_value))
                except:
                    pass
            # 根据最长单元格的内容来调整列宽
            adjusted_width = max_length + 2  # 添加一些缓冲空间
            ws.column_dimensions[column].width = adjusted_width

        print("考勤数据已成功写入 Excel 文件的 detail 表中。")

    def _get_time_or_empty(self, status_info):
        """
        获取考勤时间，如果时间不存在，返回空字符串。
        """
        return status_info["time"] if status_info["time"] else ""

