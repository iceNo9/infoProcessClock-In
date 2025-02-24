import datetime
import re
import os
import sys
import subprocess
import openpyxl
import argparse
from collections import defaultdict
from attendanceManager import AttendanceManager
from log_config import logger
from parse import *
import cmd
from utils import *
from cmdcli import *


# 工作目录
project_dir = ""


# 获取中文星期
def get_weekday_chinese(date_str):
    date = datetime.strptime(date_str, "%Y-%m-%d").date()
    weekday_num = date.weekday()  # 返回0=Monday, 1=Tuesday, ..., 6=Sunday
    weekday_chinese = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    return weekday_chinese[weekday_num]

# 生成并保存 Excel 文件
def generate_excel_file(wb, final_attendance_data):
    # 删除默认的工作表
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 创建一个工作表名为 "src"
    ws = wb.create_sheet(title="src")

    # 写入表头
    ws.append(["日期", "星期", "打卡时间1", "打卡时间2", "打卡时间3", "打卡时间4"])  # 可根据最大打卡次数调整

    # 遍历最终考勤数据
    sorted_dates = sorted(final_attendance_data.keys())
    for date in sorted_dates:
        weekday = get_weekday_chinese(date)
        formatted_times = [time.split(" ")[1] for time in final_attendance_data[date]]  # 取出时分秒部分

        # 将同一天的打卡时间写入同一行，每个时间占据一个单元格
        row = [date, weekday] + formatted_times

        # 补齐空白单元格以保证每一行的列数一致（假设最多有5次打卡）
        while len(row) < 7:  # 确保每行至少有6列（日期、星期及5个打卡时间）
            row.append('')

        ws.append(row)


def process_file(file_path, year, month, threshold_minutes, is_debug):
    global project_dir

    src_dict = convert_file(file_path)
    if(is_debug):
        save_debug_data(src_dict, project_dir, "convert")

    filter_dict = filter_times(src_dict, year, month, threshold_minutes)
    if(is_debug):
        save_debug_data(filter_dict, project_dir, "filter")

    # 写入
    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()

    # 生成 Excel 文件
    generate_excel_file(wb, filter_dict)

    # 设置输出文件路径
    output_file = file_path.replace(".txt", ".xlsx")
    
    # 假设 attendance_manager 是一个有效的对象，并调用它来处理考勤数据
    attendance_manager = AttendanceManager(year, is_flexible=True)
    result = attendance_manager.process_month(month, filter_dict)
    attendance_manager.write_attendance_to_excel(wb, result)

    wb.save(output_file)
    logger.info(f"数据已保存为：{output_file}")


def main():
    global project_dir

    MESSAGE = "支持2024.12--2025.2考勤识别;\n     新增交互界面;"
    VER = "V2.0"
    DATE = "20250224"

    # 判断是否是 Nuitka 打包环境
    is_packaged = not hasattr(sys, "_MEIPASS") and not os.path.exists(__file__)

    # 无命令行入参,进入交互命令行;否则执行命令行参数
    # 交互命令行逻辑

    logger.info(f"版本:{VER}-{DATE}")
    logger.info(f"说明:{MESSAGE}")

    # 判断是否是 Nuitka 打包环境
    is_packaged = hasattr(sys, "frozen") or not os.path.exists(__file__)

    if not is_packaged:   # 测试环境
        logger.info("运行环境:调试")
        project_dir = os.path.dirname(os.path.abspath(__file__))  # 获取当前脚本的目录        
        

        input_file_path = "test.txt"  
        year = 2025  # 使用预设年份
        month = 2  # 使用预设月份
        is_debug = True

        process_file(input_file_path, year, month, 3, is_debug)
        exit(0)

    else:
        logger.info("运行环境:打包")
        project_dir = os.path.dirname(sys.executable)  # 获取打包后的目录路径
        
    try:
        # 设置命令行参数解析器
        parser = argparse.ArgumentParser(description="处理考勤文件并生成 Excel 文件")
        parser.add_argument('file_path', nargs='?', help="考勤文件的路径")  # 可选参数
        parser.add_argument('year', nargs='?', type=int, help="年份")  # 可选参数
        parser.add_argument('month', nargs='?', type=int, help="月份")  # 可选参数
        parser.add_argument('--tm', type=int, help="设置时间阈值，单位分钟", default=3)  # 可选带参参数
        parser.add_argument('--debug', action='store_true', help="开启调试模式，传入 --debug 开启调试")

        args = parser.parse_args()

        # 如果没有命令行参数，则进入交互界面并显示帮助
        if not args.file_path:
            logger.info("进入命令行交互模式... (输入 help 获取更多命令信息)")
            IPCiCmd(project_dir).cmdloop()  # 启动交互式命令行界面
        else:
            # 如果有命令行参数，则执行文件解析
            input_file_path = args.file_path
            year = args.year
            month = args.month
            is_debug = args.debug
            threshold_minutes = args.tm


            logger.debug(f"文件路径: {input_file_path}, 年份: {year}, 月份: {month}, 过滤阈值: {threshold_minutes}, 调试模式: {is_debug}")
            
            process_file(input_file_path, year, month, threshold_minutes, is_debug)
            
    except Exception as e:
        logger.error(f"发生错误: {e}")


# 主函数，传入文件路径
if __name__ == "__main__":    
    main()

